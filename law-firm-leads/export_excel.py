#!/usr/bin/env python3
"""Export verified law firm leads to a formatted Excel workbook."""
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

leads = json.load(open("verify_results.json"))

wb = Workbook()
ws = wb.active
ws.title = "Law Firm Leads"

headers = [
    "#", "Business Name", "Phone", "Phone Verified", "Website",
    "Address", "Town", "Rating", "Reviews", "Verification Evidence",
]

# --- styles ---
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

status_colors = {
    "VERIFIED": "C6EFCE",     # green
    "NOT_FOUND": "FFEB9C",    # yellow
    "UNREACHABLE": "FFC7CE",  # red
    "NO_WEBSITE": "D9D9D9",   # gray
}

# header row
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border

# data rows
for i, lead in enumerate(leads, 1):
    row = i + 1
    values = [
        i,
        lead.get("name"),
        lead.get("phone"),
        lead.get("phone_verified"),
        lead.get("website"),
        lead.get("address"),
        lead.get("town_searched"),
        lead.get("rating"),
        lead.get("reviews"),
        lead.get("verification_evidence"),
    ]
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = border
        c.alignment = center if col in (1, 4, 7, 8, 9) else left
        if col == 4:  # color-code verification status
            fill = status_colors.get(v)
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
                c.font = Font(bold=True)

# column widths
widths = [4, 32, 16, 14, 40, 38, 12, 8, 9, 28]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads)+1}"

# summary sheet
s = wb.create_sheet("Summary")
verified = sum(1 for l in leads if l["phone_verified"] == "VERIFIED")
s["A1"] = "Law Firm Leads - Westport / Greenwich, CT"
s["A1"].font = Font(bold=True, size=14)
rows = [
    ("Total leads", len(leads)),
    ("Phones VERIFIED", verified),
    ("Not found on site", sum(1 for l in leads if l["phone_verified"] == "NOT_FOUND")),
    ("Site unreachable", sum(1 for l in leads if l["phone_verified"] == "UNREACHABLE")),
    ("Verification rate", f"{verified/len(leads)*100:.0f}%"),
    ("Source", "Apify Google Maps Scraper"),
]
for i, (k, v) in enumerate(rows, 3):
    s.cell(row=i, column=1, value=k).font = Font(bold=True)
    s.cell(row=i, column=2, value=v)
s.column_dimensions["A"].width = 22
s.column_dimensions["B"].width = 30

wb.save("Law_Firm_Leads_CT.xlsx")
print("Saved Law_Firm_Leads_CT.xlsx")
